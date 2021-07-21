from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import openpyxl as xl
# import pyautogui
import time

options = webdriver.ChromeOptions()
options.add_argument(r"user-data-dir=C:\Users\Mitul\AppData\Local\Google\Chrome\Selenium")
options.headless = False

PATH = r'C:\chromedriver\chromedriver.exe'

driver = webdriver.Chrome(executable_path=PATH, options=options)
driver.get('https://web.whatsapp.com/')

wb = xl.load_workbook(r'C:\Users\Mitul\Desktop\Contacts.xlsx')
sheet = wb['Sheet1']

choice = input("Do you want to send text message or media files[T/M]: ")
count = 1

for row in range(2, sheet.max_row + 1):
    name = sheet.cell(row, 1)
    name = name.value
    print(name)

    # name = input('Enter the name of the person or the group or the broadcast you want to send message to: ')

    if choice.upper() == 'T':
        message = sheet.cell(row, 2)
        # message = message.value
        message = f'''Hello {name}! 
        
{message.value}'''
        print(message)
        if count == 1:
            # message = input('Enter the message you want to send: ')
            input('Press any key after scanning the QR code ONLY ')
            count += 1

        searchbox = driver.find_element_by_xpath('//*[@id="side"]/div[1]/div/label/div/div[2]')
        searchbox.click()
        searchbox.send_keys(name)
        searchbox.send_keys(Keys.RETURN)

        textbox = driver.find_element_by_xpath('//*[@id="main"]/footer/div[1]/div[2]/div/div[2]')
        textbox.click()

        textbox.send_keys(message)
        driver.implicitly_wait(30)
        textbox.send_keys(Keys.RETURN)

    elif choice.upper() == 'M':
        if count == 1:
            type_choice = input(''' What do you want to send?
            For any attachments : Press 'P'
            Photos or Videos from Camera: Press 'C'
            Room: 'R'
            ''')
            input('Press any key after scanning the QR code ONLY ')
            count += 1
        # Documents: Press 'D'
        # Contact: Press 'Ct'

        # time.sleep(5)
        media_path = sheet.cell(row, 3)
        media_path = media_path.value

        searchbox = driver.find_element_by_xpath('//*[@id="side"]/div[1]/div/label/div/div[2]')
        searchbox.click()
        searchbox.send_keys(name)
        searchbox.send_keys(Keys.RETURN)

        attach_pin = driver.find_element_by_xpath('//div[@title="Attach"]')
        attach_pin.click()

        if type_choice.upper() == 'P':
            # media_path = input('Enter the exact path to the photo or video file you want to send: ')
            # input('Press any key ONLY after scanning the QR code ')
            image_button = driver.find_element_by_xpath(
                '//*[@id="main"]/footer/div[1]/div[1]/div[2]/div/span/div[1]/div/ul/li[1]/button/input')
            image_button.send_keys(media_path)
            time.sleep(3)
            media_send_btn = driver.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[2]/div[2]/span/div[1]/span/div[1]/div/div[2]/span/div/div/span')
            media_send_btn.click()

            # image_button.click()
            # time.sleep(1)
            # pyautogui.typewrite(media_path)
            # pyautogui.press('enter')
            # time.sleep(2)
            # pyautogui.press('enter')

        elif type_choice.upper() == 'C':
            camera_button = driver.find_element_by_xpath(
                '//*[@id="main"]/footer/div[1]/div[1]/div[2]/div/span/div[1]/div/ul/li[2]/button/span')
            camera_button.click()

        # elif type_choice.upper() == 'D':
        #     media_path = input('Enter the name of the contact you want to send: ')
        #     input('Press any key ONLY after scanning the QR code ')
        #     document_button = driver.find_element_by_xpath(
        #         '//*[@id="main"]/footer/div[1]/div[1]/div[2]/div/span/div[1]/div/ul/li[3]/button/span')
        #     # document_button.send_keys(media_path)
        #     document_button.click()
        #     time.sleep(1)
        #     pyautogui.typewrite(media_path)
        #     pyautogui.press('enter')
        #     time.sleep(2)
        #     pyautogui.press('enter')

        # elif type_choice.upper() == 'CT':
        #     media_path = input('Enter the name of the contact you want to send: ')
        #     input('Press any key ONLY after scanning the QR code ')
        #     contact_button = driver.find_element_by_xpath(
        #         '//*[@id="main"]/footer/div[1]/div[1]/div[2]/div/span/div[1]/div/ul/li[4]/button/span')
        #     contact_button.click()
        #     contact_search = driver.find_element_by_xpath(
        #         '//*[@id="app"]/div[1]/span[2]/div[1]/span/div[1]/div/div/div/div/div/div[1]/div/label/div/div[2]')
        #     contact_search.send_keys(media_path)
        #     contact_search.send_keys(Keys.RETURN)
        #     contact_send = driver.find_element_by_xpath(
        #         '//*[@id="app"]/div[1]/span[2]/div[1]/span/div[1]/div/div/div/div/div/span/div/div/div/span')
        #     contact_send.click()

        elif type_choice.upper() == 'R':
            input('Press any key ONLY after scanning the QR code ')
            room_button = driver.find_element_by_xpath(
                '//*[@id="main"]/footer/div[1]/div[1]/div[2]/div/span/div[1]/div/ul/li[5]/button/span')
            room_button.click()
            messanger_button = driver.find_element_by_xpath(
                '//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div[2]/div[2]/div/div')
            messanger_button.click()
